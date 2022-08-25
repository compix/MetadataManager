import typing
from table.Table import Table
from table.ExcelTable import ExcelTable
from table.CSVTable import CSVTable
import os

try:
    import xlrd
except:
    pass

try:
    import openpyxl
except:
    pass

def readTable(tablePath: str, csvSeparator=';', excelSheetName=None, encodingOverride=None) -> Table:
    if tablePath.lower().endswith('.csv'):
        return CSVTable(tablePath, csvSeparator, encodingOverride=encodingOverride)
    elif tablePath.lower().endswith('.xlsx') or tablePath.lower().endswith('.xls'):
        return ExcelTable.read(tablePath, excelSheetName, encodingOverride=encodingOverride)
    else:
        raise RuntimeError(f'Unsupported file format {os.path.splitext(tablePath)[1]}')

def getSheetNames(tablePath: str):
    if tablePath.lower().endswith('.xls'):
        wb = xlrd.open_workbook(filename=tablePath)
        return wb.sheet_names()
    elif tablePath.lower().endswith('.xlsx'):
        wb = openpyxl.load_workbook(tablePath, read_only=True, data_only=True)
        sheetnames = wb.sheetnames
        wb.close()
        
        return sheetnames

    return None

def zipRowAndHeader(row: typing.List[str], header: typing.List[str], rowIndices: typing.List[str]):
    return {header[hi]: row[i] for hi, i in enumerate(rowIndices) if i < len(row)}