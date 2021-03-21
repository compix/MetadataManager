from table.Table import Table

imported = False
try:
    import xlrd
    imported = True
except:
    print("Note: Reading xls files is not available because the xlrd module is missing - pip install xlrd")

try:
    import openpyxl
    import io
except:
    print("Note: Reading xlsx files is not available because the openpyxl module is missing - pip install xlrd")

class ExcelTable(Table):
    def __init__(self, sheet):
        super().__init__()

        self.sheet = sheet

    def getRowValues(self, rowIndex):
        pass

    def getColumnValues(self, colIndex):
        pass

    def getCellValue(self, rowIndex, colIndex):
        pass

    @property
    def ncols(self):
        pass

    @property
    def nrows(self):
        pass

    @staticmethod
    def read(workbookPath: str, sheetName: str, encodingOverride=None):
        if workbookPath.lower().endswith('.xls'):
            return XlsTable(xlrd.open_workbook(filename=workbookPath,encoding_override=encodingOverride).sheet_by_name(sheetName))
        else:
            with open(workbookPath, "rb") as f:
                inMemFile = io.BytesIO(f.read())

            wb = openpyxl.load_workbook(inMemFile, read_only=True, data_only=True)
            return XlsxTable(wb.get_sheet_by_name(sheetName))

class XlsTable(ExcelTable):
    def getRowValues(self, rowIndex):
        return self.sheet.row_values(rowIndex) if self.sheet != None else []

    def getColumnValues(self, colIndex):
        return self.sheet.col_values(colIndex) if self.sheet != None else []

    def getCellValue(self, rowIndex, colIndex):
        return self.sheet.cell_value(rowIndex, colIndex) if self.sheet != None else None

    @property
    def ncols(self):
        return self.sheet.ncols if self.sheet != None else 0

    @property
    def nrows(self):
        return self.sheet.nrows if self.sheet != None else 0

class XlsxTable(ExcelTable):
    def getRowValues(self, rowIndex):
        return [cell.value if cell else None for row in self.sheet.iter_rows(min_row=rowIndex+1, max_row=rowIndex+1) for cell in row] if self.sheet != None else []

    def getColumnValues(self, colIndex):
        return [cell.value if cell else None for col in self.sheet.iter_cols(min_col=colIndex+1, max_col=colIndex+1) for cell in col] if self.sheet != None else []

    def getCellValue(self, rowIndex, colIndex):
        cell = self.sheet.cell(rowIndex, colIndex) if self.sheet != None else None
        if cell:
            return cell.value
        
        return None

    @property
    def ncols(self):
        return self.sheet.max_column if self.sheet != None else 0

    @property
    def nrows(self):
        return self.sheet.max_row if self.sheet != None else 0